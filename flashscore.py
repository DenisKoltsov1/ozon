from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from fake_useragent import UserAgent
from selenium.webdriver.common.keys import Keys
import undetected_chromedriver 
import os
options = webdriver.ChromeOptions()
from tkinter import *
from tkinter import ttk
import pandas as pd
import glob
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


name=" "
price=" "
priceSkidka=" "
nds=" "
KomercialType=" "
brendItem=" "
typeItem=" "
obiem=" "
articul= " "
width=" "

hidth=" "
anot=" "
dlina=" "
femal=" "
ManufacturerСountry=" "
linkUrl=" "
count= " "
Sostav= " "
keyWord=[]
linkUrldop=[]
linkUrldop1=""
aromaClass=" "
keyWord1=" "
#name,price,priceSkidka,nds,KomercialType

def parser(id):
    driver=undetected_chromedriver.Chrome()
    
    global name, price,priceSkidka,nds,KomercialType,brendItem,typeItem,obiem,countItem,femal,classop,ManufacturerСountry,articul,width,linkUrl,hidth,anot,dlina,keyWord,linkUrldop,aromaClass,Sostav,keyWord1,linkUrldop1   
    
    time.sleep(1)
    # переход  на сайт  ozon
    html=driver.get("https://www.ozon.ru/")
    time.sleep(1)
    #ищем  поле поиска
    kcal = driver.find_element(By.XPATH,"//input[contains(@class,'tsBodyL')]")
    time.sleep(1)
    #вставляем  id
    kcal.send_keys(id)
    time.sleep(1)
    #лупа для продолжения поиска
    #lupa = driver.find_element(By.XPATH,'//button[@class="x5-a1 x5-f0"]')
    #header-search-go
    lupa = driver.find_element(By.XPATH,'//button[@type="submit"]')
    time.sleep(1)
    lupa.click()

    time.sleep(1)

    # клик по ссылке  продукта 
    item=driver.find_element(By.XPATH,"//span[contains(@class,'tsBodyL')]")
    item.click()
    time.sleep(1)
    

    
    
    # ищем имя  товара
    
    try: 
        #name=driver.find_element(By.XPATH,"//h1[contains(@class,'rn')]").text
        name= driver.find_element(By.XPATH,"//div[@data-widget='webProductHeading']").text
        name=name
        print(name)
    
    except:
        name=''
        print('элемент name(имя) не найден')     
    
    #ищем цену  товара
    try:
        #price=driver.find_element(By.XPATH,"//span[contains(@class,'p1n np2')]").text.replace("₽", "")
        price= driver.find_element(By.XPATH,"//div[@data-widget='webPrice']//child::span[1]").text.replace("₽", "")
        price=price
        print(price)
        
    except:
        price=''
        print('элемент price(цена) не найден')
    
    #ищем цену  до скидки товара
    try:
        #priceSkidka=driver.find_element(By.XPATH,"//span[contains(@class,'pn2')]").text.replace("₽", "")
        priceSkidka=driver.find_element(By.XPATH,"//div[@data-widget='webPrice']//child::span[2]").text.replace("₽", "") 
        priceSkidka=priceSkidka 
        print(priceSkidka)      
    except:
        priceSkidka=''
        print('элемент priceSkidka(цена со скидкой) не найден')   
          
    #ищем  nds
    nds='13%'
    nds=nds
    print(nds)
 

    #KomercialType
    try:
        KomercialType=driver.find_element(By.XPATH,"//span[text()='Тип']/following::dd[1]").text
        KomercialType=KomercialType
        print(KomercialType)
    except:
        KomercialType=''
        print('элемент KomercialType не найден')
    
    #Ширина
    try:
        
        width=driver.find_element(By.XPATH,"//span[text()='Размер упаковки (Длина х Ширина х Высота), см']/following::dd[1]").text
        #r=re.search('[0-9]',width)
        r=re.search(r"\d{1,3}",width)
        width=r[0]
        print(r[0])
    except:
        width=''
        print('элемент width не найден')
    #высота
    try:
        #r=re.match(r'\d[0-9]')
        hidth=driver.find_element(By.XPATH,"//span[text()='Размер упаковки (Длина х Ширина х Высота), см']/following::dd[1]").text
        r2=re.search(r"[*]\d{1,3}",hidth)
        hidth=r2[0].replace("*", "")
        print(r2[0].replace("*", ""))
    except:
        hidth=''
        print('элемент hidth не найден')
    #Длина
    try:
        dlina=driver.find_element(By.XPATH,"//span[text()='Размер упаковки (Длина х Ширина х Высота), см']/following::dd[1]").text
        r3=re.search(r"\d{1,3}$",dlina)
        dlina=r3[0]
        print(r3[0])
    except:
        width=''
        print('элемент width не найден')

    #url ссылка на картинку
    try:

        linkUrl=driver.find_element(By.XPATH,'//img[@fetchpriority="high"]').get_attribute("src")

        
        linkUrl=linkUrl
        print(linkUrl)
    except:
        linkUrl=''
        print('элемент LinkURL не найден')
 
 
    #url ссылка на доп картинку
    try:
        
        linkUrldop.append(driver.find_element(By.XPATH,'//div[@data-index="1"]/child::div/child::img').get_attribute("src"))
        linkUrldop.append(driver.find_element(By.XPATH,'//div[@data-index="2"]/child::div/child::img').get_attribute("src"))
        linkUrldop.append(driver.find_element(By.XPATH,'//div[@data-index="3"]/child::div/child::img').get_attribute("src"))
        linkUrldop.append(driver.find_element(By.XPATH,'//div[@data-index="4"]/child::div/child::img').get_attribute("src"))

        linkUrldop1=",".join(linkUrldop)
        linkUrldop1=linkUrldop1.replace(",",";")
        
        print(linkUrldop1)   
        
    except:
        linkUrldop=''
        print('элемент linkUrldop не найден')

    #SerialNumber
    #SerialNumber=driver.find_element(By.XPATH,"//span[text()='Тип']/following::dd[1]")
    #print(SerialNumber.text)
    #ищем  id 
    Ozonid=id
    print(Ozonid)
    #ищем тип  товара 
    
    #бренд   товара
    try:
        brendItem=driver.find_element(By.XPATH,"//span[text()='Бренд']/following::dd[1]").text
        brendItem= brendItem
        print(brendItem)
    except:
        brendIte=''
        print('элемент brendItem не найден')          
    time.sleep(45)

    #тип  товара
    #/td/span[text()='Данные2']/following::td[1]/span
    try:
        typeItem=driver.find_element(By.XPATH,"//span[text()='Тип']/following::dd[1]").text
        typeItem=typeItem
        print(typeItem)
        type(typeItem)
    except:
        typeItem=''
        print('элемент typeItem не найден')

    #объем товара
    try:
        obiem=driver.find_element(By.XPATH,"//span[text()='Объем, мл']/following::dd[1]").text
        print(obiem)
    except:
        obiem=''
        print('элемент obiem не найден')
    #едениц товара
    countItem=1
    countItem=countItem
    print(countItem)  
    
    #артикул товара
    try:
        articul=driver.find_element(By.XPATH,"//span[text()='Артикул']/following::dd[1]").text
        articul=articul
        print(articul)
    except:
        articul=''
        print('элемент articul не найден')
    
    #едениц товара
    countItem=1
    countItem=countItem
    print(countItem)  

    #пол
    try:
        femal=driver.find_element(By.XPATH,"//span[text()='Целевая аудитория']/following::dd[1]").text
        femal=femal
        print(femal)
    except:
        femal=''  
        print('Элемент femal не найден')


    #аннотации
    try:
        #anot=driver.find_element_by_id("section-description")/child::div[4]")
        anot=driver.find_element(By.XPATH,"//h2[text()='Описание']/following::div[1]").text

        anot=anot
        print(anot)
    except:
        anot=''  
        print('Элемент anot не найден')


    #страна изготовитель
    try:
        ManufacturerСountry=driver.find_element(By.XPATH,"//span[text()='Страна-изготовитель']/following::dd[1]").text
        ManufacturerСountry=ManufacturerСountry
        print(ManufacturerСountry)
    except:
        ManufacturerСountry=''
        print('элемент ManufactureCountry не найден')    
     


    #классификация аромата
    try:
        aromaClass=driver.find_element(By.XPATH,"//span[text()='Классификация аромата']/following::dd[1]").text
        aromaClass=aromaClass
        print(aromaClass)
    except:
        aromaClass=''
        print('элемент aromaClass  не найден') 
    #состав
    try:
        Sostav=driver.find_element(By.XPATH,"//*[text()='Состав']/following-sibling::p").text
        Sostav=Sostav
        Sostav=Sostav.replace(", ",";")
        
        print(Sostav)
    except:
        Sostav=''
        print('элемент Sostav  не найден')    
   
   #ключевые  слова
    
    try:
        #,aromaClass
        keyWord=[brendItem,aromaClass,Sostav,typeItem]
        
        keyWord1=','.join(keyWord)

        keyWord1=keyWord1.replace(",", ";")

        #keyWord1=keyWord1.replace("; ;", "")   
        print(keyWord1)
    except:
        keyWord1=''
        print("элемент keyWord не найден")


    driver.close()  
    driver.quit()

    return id,name,price,priceSkidka,nds,KomercialType,articul
 
 
    

#,name,price,priceSkidka,nds,KomercialType
#articul,name,price,priceSkidka,nds,KomercialType,obiem,width,higth,dlina,linurl,linkurldrop,brendItem,femal,anot,Sostav,CoutItem
def OpenPixx(file,id,r):
    hel='hello'
    wb = load_workbook(file)
    # get_sheet_names() - выводит список с названием листов,
    ws = wb['Шаблон для поставщика']
    #ws['B4']=id
    
    #i=1
    ws.cell(row=r, column=1).value = count  #номер  столбца
    ws.cell(row=r, column=2).value = articul    #артикул
    ws.cell(row=r, column=3).value = name   #название товара
    ws.cell(row=r, column=4).value = price  #цена товара
    ws.cell(row=r, column=5).value = priceSkidka # цена до скидки
    ws.cell(row=r, column=6).value = None #nds
    ws.cell(row=r, column=7).value = id     #id 7 поле
    ws.cell(row=r, column=8).value = KomercialType
    ws.cell(row=r, column=9).value = None
    ws.cell(row=r, column=10).value = obiem   #вес в упаковке в г 10 поле
    ws.cell(row=r, column=11).value = width #ширина упаковки
    ws.cell(row=r, column=12).value = hidth #высота  упаковки
    ws.cell(row=r, column=13).value = dlina #длина упаковки 13 поле
    ws.cell(row=r, column=14).value = linkUrl #cсылка на главное фото
    ws.cell(row=r, column=15).value = linkUrldop1 #ссылка на доп фото      
    ws.cell(row=r, column=16).value = None #ссылка на фото 360
    ws.cell(row=r, column=17).value = None #артикул фото
    ws.cell(row=r, column=18).value = brendItem #бренд 18 поле
    ws.cell(row=r, column=19).value = typeItem #тип
    ws.cell(row=r, column=20).value = None #Объединить на одной карточке
    ws.cell(row=r, column=21).value = obiem #Объем, мл*
    ws.cell(row=r, column=22).value = countItem #Единиц в одном товаре
    ws.cell(row=r, column=23).value = femal #пол
    ws.cell(row=r, column=24).value = None #Срок годности в днях
    ws.cell(row=r, column=25).value = None #Класс опасности товара
    ws.cell(row=r, column=26).value = None #Rich-контент JSON
    ws.cell(row=r, column=27).value = None #Минимальный возраст ребенка
    ws.cell(row=r, column=28).value = None #Максимальный возраст ребенка
    ws.cell(row=r, column=29).value = None #Планирую доставлять товар в нескольких упаковках
    ws.cell(row=r, column=30).value = anot #Аннотация
    
    ws.cell(row=r, column=31).value = keyWord1 #Ключевые слова
    ws.cell(row=r, column=32).value = None #Серии
    ws.cell(row=r, column=33).value = aromaClass #Классификация аромата
    ws.cell(row=r, column=34).value = ManufacturerСountry #Страна-изготовитель
    ws.cell(row=r, column=35).value = obiem #Вес, г
    ws.cell(row=r, column=36).value = None #Упаковка
    ws.cell(row=r, column=37).value = None #Целевая аудитория
    ws.cell(row=r, column=38).value = Sostav #Состав
    ws.cell(row=r, column=39).value = None #Особенности состава
    ws.cell(row=r, column=40).value = None #Вид парфюмерии (служебный)
    ws.cell(row=r, column=41).value = None #Код ОКПД/ТН ВЭД для парфюмерии
    ws.cell(row=r, column=42).value = None #Количество заводских упаковок
    ws.cell(row=r, column=43).value = None #Ошибка
    ws.cell(row=r, column=44).value = None #Предупреждение
    
    wb.save(file)
    return 1


def  get_str_number(file):
    wb = load_workbook(file)
    ws = wb['Шаблон для поставщика']
    i=4
    count=1 
    if (ws.cell(row=i,column=1).value) ==None:
        r=4
        count=1
    else:
        for i  in  range(1,1000):
            if (ws.cell(row=i,column=1).value) !=None:
                r=ws.cell(row=i,column=1).row
                #count=ws.cell(row=i,column=1).value
                r+=1
                
                
    #print(r)
    wb.close()
    return r           

def get_count(r):
    global count
    count=r-3                
    #print(count)
    
    return count           
 
id=154369096
'''
parser(id)
'''
'''
file ="C:\\ozon1\\Парфюмерия.xlsx"
r=get_str_number(file)
count=get_count(r)
OpenPixx(file,id,r)

'''